import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_styled_excel(table_widget, file_name):
    # Créer un nouveau fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données"

    # Définir un style de bordure pour les cellules
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Définir une couleur de fond pour les en-têtes
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Définir une couleur de fond pour les lignes impaires
    row_fill_odd = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    # Définir les en-têtes de colonnes
    headers = ['CNE', 'Date', 'Status', 'Operator']
    
    # Appliquer les en-têtes avec un fond coloré et en gras
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")  # Texte en gras et blanc
        cell.fill = header_fill  # Fond coloré pour l'en-tête
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrer le texte
        cell.border = border_style  # Appliquer la bordure
    
    # Récupérer les données de la table et les ajouter à l'Excel
    for row_num in range(table_widget.rowCount()):
        for col_num in range(table_widget.columnCount()):
            item = table_widget.item(row_num, col_num)
            if item is not None:
                # Déterminer la couleur de fond pour chaque ligne (alternance des couleurs)
                if row_num % 2 == 0:
                    row_fill = row_fill_odd
                else:
                    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                # Remplir la cellule avec les données
                cell = ws.cell(row=row_num + 2, column=col_num + 1, value=item.text())
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrer le texte
                cell.border = border_style  # Appliquer la bordure
                cell.fill = row_fill  # Appliquer la couleur de fond de la ligne

    # Ajuster la largeur des colonnes
    for col_num in range(len(headers)):
        column = get_column_letter(col_num + 1)
        max_length = 0
        for row in ws.iter_rows(min_col=col_num + 1, max_col=col_num + 1):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Enregistrer le fichier Excel
    wb.save(f"csv/{file_name}.xlsx")
    print(f"Le fichier Excel a été généré avec succès : {file_name}")

# Exemple d'utilisation
# Assurez-vous que votre tableWidget_m est déjà rempli avec des données avant de l'utiliser
# generate_styled_excel(self.tableWidget_m, "styled_output.xlsx")
