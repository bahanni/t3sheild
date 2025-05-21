import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_styled_excel(flat, finale, date_str, file_name):
    """
    Crée un Excel qui :
     - Affiche date_str une seule fois en haut.
     - Présente deux matrices (flat et finale) dans leur forme (rows × cols).
    """
    # Créer le dossier de sortie si nécessaire
    os.makedirs("./csv", exist_ok=True)

    # Nouveau classeur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrice"

    # Styles
    border = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )
    header_fill = PatternFill("solid", fgColor="4F81BD")
    odd_fill    = PatternFill("solid", fgColor="D9EAF7")
    even_fill   = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal='center', vertical='center')

    # 1) Date en A1/B1
    ws["A1"] = "Date :"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = date_str
    ws["B1"].alignment = center

    # Déterminer dimensions à partir de 'finale'
    rows = len(finale)
    cols = len(finale[0]) if rows else 0

    # Fonction utilitaire pour écrire une matrice
    def _write_matrix(start_row, title, matrix):
        # Titre
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = Font(bold=True)
        # En-têtes de colonnes
        for c in range(cols):
            hdr = ws.cell(row=start_row+1, column=2+c, value=f"Col {c+1}")
            hdr.font = Font(bold=True, color="FFFFFF")
            hdr.fill = header_fill
            hdr.alignment = center
            hdr.border = border
        # En-têtes de lignes + données
        for r in range(rows):
            # en-tête de ligne
            hdr = ws.cell(row=start_row+2+r, column=1, value=f"Ligne {r+1}")
            hdr.font = Font(bold=True, color="FFFFFF")
            hdr.fill = header_fill
            hdr.alignment = center
            hdr.border = border
            # données
            for c in range(cols):
                val = matrix[r][c]
                cell = ws.cell(row=start_row+2+r, column=2+c, value=str(val))
                cell.alignment = center
                cell.border = border
                cell.fill = odd_fill if (r % 2 == 0) else even_fill

    # 2) Matrice 'flat' : on la reconstruit en 2D row-major
    flat2d = [
        flat[r*cols:(r+1)*cols]
        for r in range(rows)
    ]
    _write_matrix(start_row=3, title="Matrice Flat", matrix=flat2d)

    # 3) Matrice 'finale' : affichée juste en-dessous
    second_block_row = 3 + 2 + rows + 1
    _write_matrix(start_row=second_block_row, title="Matrice Finale", matrix=finale)

    # 4) Ajuster largeur des colonnes
    for col_idx in range(1, cols + 2):  # +2 pour la colonne d'en-têtes
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value))
            for r in range(1, ws.max_row+1)
        )
        ws.column_dimensions[letter].width = max_len + 2

    # Sauver
    path = f"./csv/{file_name}.xlsx"
    wb.save(path)
    print(f"Excel généré : {path}")


if __name__ == "__main__":
    # Exemple de matrice 4×3
    finale = [
        [('L','11',-55), ('L','21',-55), ('L','31',-55)],
        [('L','12',-55), ('L','22',-55), ('L','32',-55)],
        [('L','13',-55), ('L','23',-55), ('L','33',-55)],
        [('L','14',-55), ('L','24',-55), ('L','34',-55)]
    ]
    # Aplatir en row-major
    flat = [elem for row in finale for elem in row]
    # Date unique
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    generate_styled_excel(flat, finale, date_str, "test_matrices")
